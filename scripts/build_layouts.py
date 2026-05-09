"""
Extract per-file fixed-width layouts from each release's Data_Layout XLSX.

Outputs one CSV per data file into {out_dir}/layout/{file_key}_layout.csv,
each row = one field with srl, name, block, item, length, byte_start,
byte_end, field_name, remarks.

Usage:
    python3 scripts/build_layouts.py                # all known releases
    python3 scripts/build_layouts.py annual_2023_24 # single release
"""

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from releases import RELEASES, ROOT
from canonicalize import canonical_name, slug


def extract_data_layout_section(ws, start_row, end_row):
    """Yield rows from 'Data Layout' sheet between section bounds.

    Skips file-header rows and the column-header row. Accepts rows where Srl
    is missing as long as length + byte positions are populated — some PLFS
    layouts (e.g., 2018-19 perrv) have srl-less footer rows for the generated
    weight columns.
    """
    fallback_srl = 0
    for r in range(start_row, end_row + 1):
        srl = ws.cell(r, 1).value
        length = ws.cell(r, 5).value
        bs = ws.cell(r, 6).value
        be = ws.cell(r, 7).value
        is_data_row = (
            isinstance(srl, int)
            or (
                isinstance(length, int) and length > 0
                and isinstance(bs, int) and isinstance(be, int)
            )
        )
        if not is_data_row:
            continue
        if not isinstance(srl, int):
            srl = fallback_srl + 1
        fallback_srl = srl
        yield {
            "srl": srl,
            "name": ws.cell(r, 2).value,
            "block": ws.cell(r, 3).value,
            "item": ws.cell(r, 4).value,
            "length": length,
            "byte_start": bs,
            "byte_end": be,
            "remarks": ws.cell(r, 8).value,
        }


def extract_field_names_by_srl(ws, fieldname_col):
    """Map srl -> field_name from a per-file sheet (Srl is in col A)."""
    out = {}
    for r in range(1, ws.max_row + 1):
        srl = ws.cell(r, 1).value
        if not isinstance(srl, int):
            continue
        out[srl] = ws.cell(r, fieldname_col).value
    return out


def extract_field_names_in_order(ws, fieldname_col, length_col=4):
    """Return [field_name, ...] from a per-file sheet that has no Srl column.

    Skips header/title rows (anything where length_col isn't a positive int).
    """
    out = []
    for r in range(1, ws.max_row + 1):
        L = ws.cell(r, length_col).value
        if not (isinstance(L, int) and L > 0):
            continue
        out.append(ws.cell(r, fieldname_col).value)
    return out


def extract_self_contained(ws, byte_start_col, byte_end_col, fieldname_col):
    """Pull the entire layout from a per-file sheet that has byte positions."""
    rows = []
    for r in range(1, ws.max_row + 1):
        srl = ws.cell(r, 1).value
        if not isinstance(srl, int):
            continue
        rows.append(
            {
                "srl": srl,
                "field_name": ws.cell(r, fieldname_col).value,
                "name": ws.cell(r, 2).value,
                "block": ws.cell(r, 3).value,
                "item": ws.cell(r, 4).value,
                "length": ws.cell(r, 5).value,
                "byte_start": ws.cell(r, byte_start_col).value,
                "byte_end": ws.cell(r, byte_end_col).value,
                "remarks": ws.cell(r, 9).value if ws.max_column >= 9 else "",
            }
        )
    return rows


def write_layout(out_path: Path, rows):
    """Write the layout CSV. Field names are normalised to lower-case so column
    names are consistent across releases (the source XLSX is mixed case)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "srl", "field_name", "name", "block", "item",
                "length", "byte_start", "byte_end", "remarks",
            ]
        )
        total = 0
        for r in rows:
            fn = (r.get("field_name") or "").strip().lower()
            w.writerow([
                r["srl"], fn, r["name"],
                r["block"] or "", r["item"] or "",
                r["length"], r["byte_start"], r["byte_end"],
                r["remarks"] or "",
            ])
            total += r["length"] or 0
    return total, len(list(rows))


def write_state_codes(wb, out_path: Path):
    """Pull the State code sheet (if present) into codemaps/state.csv."""
    sheet_names = [s.lower() for s in wb.sheetnames]
    candidates = ("state code", "state codes", "state")
    for cand in candidates:
        if cand in sheet_names:
            ws = wb[wb.sheetnames[sheet_names.index(cand)]]
            with out_path.open("w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["state_code", "state_name"])
                for r in range(1, ws.max_row + 1):
                    code = ws.cell(r, 1).value
                    name = ws.cell(r, 2).value
                    if code is None or name is None:
                        continue
                    code_s = str(code).strip()
                    if not code_s.isdigit():
                        continue
                    w.writerow([code_s.zfill(2), str(name).strip()])
            return True
    return False


def _disambiguate_canonical(rows):
    """Resolve per-day CWS columns and other repeated-name fields by Item code.

    Block 6 has fields like 'Industry Code (NIC) for activity 1' / 'for activity 2'
    repeated for items 5/3.1 ... 5/3.7 (7 days × 2 activities). We assign
    mnemonics like ind11 (activity 1, day 1), ind21 (activity 2, day 1),
    ind12 (activity 1, day 2), etc., matching CY2024's convention.
    """
    used = {}
    for r in rows:
        fn = r.get("field_name")
        if not fn:
            continue
        # Detect the per-day CWS pattern: name contains "for activity 1" / "2"
        name = (r.get("name") or "").lower()
        item = str(r.get("item") or "")
        block = str(r.get("block") or "")
        # Only block 6, item like '5/3.X' triggers serial naming
        if block == "6" and "/" in item:
            try:
                day_part = item.split("/")[1]   # e.g., '3.1'
                day = int(day_part.split(".")[1])
            except (IndexError, ValueError):
                day = None
            if day is not None:
                activity = 1 if "activity 1" in name else (2 if "activity 2" in name else None)
                # Per-activity per-day fields
                if activity and "industry" in name:
                    r["field_name"] = f"ind{activity}{day}"
                elif activity and "occupation" in name:
                    r["field_name"] = f"ocu{activity}{day}"
                elif activity and "status code" in name:
                    r["field_name"] = f"sts{activity}{day}"
                elif activity and "wage earning" in name:
                    r["field_name"] = f"ern{activity}{day}"
                elif activity and "hours" in name and "actu" in name:
                    r["field_name"] = f"hrs{activity}{day}"
                # Per-day-only fields (no activity index)
                elif "total hours" in name:
                    r["field_name"] = f"tot_hrs{day}"
                elif "hours avail" in name or "hours available" in name:
                    r["field_name"] = f"hav{day}"
                elif "duration of engagement" in name:
                    r["field_name"] = f"dur_eng{day}"

        # Block 5.3: Duration of engagement (Principal vs Subsidiary)
        if block == "5.3" and "duration of engagement" in name:
            if "principal" in name:
                r["field_name"] = "dur_pas"
            elif "subsidiary" in name:
                r["field_name"] = "dur_sas"

        # Block 6: in older layouts both ern_reg & ern_self share the same Full Name
        # (PLFS source typo). Differentiate by Item: 9=ern_reg, 10=ern_self.
        if block == "6" and "earnings" in name and "regular" in name:
            item_str = str(r.get("item") or "")
            if item_str in ("9",): r["field_name"] = "ern_reg"
            elif item_str in ("10",): r["field_name"] = "ern_self"

        # Track for collision detection
        used[r["field_name"]] = used.get(r["field_name"], 0) + 1


def build_release(release_name: str):
    cfg = RELEASES[release_name]
    print(f"\n=== {release_name} — {cfg['label']} ===")
    wb = openpyxl.load_workbook(cfg["xlsx"], data_only=True)

    summary = []
    is_csv_input = cfg.get("input_kind") in ("csv", "tsv")
    for fc in cfg["files"]:
        key = fc["key"]
        if is_csv_input:
            # Read Data Layout section and apply canonicalize() to derive field_name
            layout_ws = wb["Data Layout"]
            rows = list(extract_data_layout_section(layout_ws, *fc["section"]))
            for r in rows:
                cn = canonical_name(r["name"], r.get("block"), r.get("item"))
                r["field_name"] = cn or slug(r["name"])
            _disambiguate_canonical(rows)
        elif cfg.get("self_contained"):
            # CY2025 mode: per-file sheet has byte_start/byte_end + field_name directly
            ws = wb[fc["fieldname_sheet"]]
            rows = extract_self_contained(
                ws,
                byte_start_col=fc["byte_start_col"],
                byte_end_col=fc["byte_end_col"],
                fieldname_col=fc["fieldname_col"],
            )
        else:
            layout_ws = wb["Data Layout"]
            section_rows = list(
                extract_data_layout_section(layout_ws, *fc["section"])
            )
            ws = wb[fc["fieldname_sheet"]]
            if fc["srl_join"]:
                names = extract_field_names_by_srl(ws, fc["fieldname_col"])
                for r in section_rows:
                    r["field_name"] = names.get(r["srl"], "")
            else:
                ordered = extract_field_names_in_order(
                    ws, fc["fieldname_col"]
                )
                if len(ordered) != len(section_rows):
                    print(
                        f"  WARN {key}: ordinal merge size mismatch "
                        f"(layout={len(section_rows)} fields={len(ordered)})"
                    )
                for r, fn in zip(section_rows, ordered):
                    r["field_name"] = fn
            rows = section_rows

        out_path = cfg["layout_dir"] / f"{key}_layout.csv"
        total_bytes, n = write_layout(out_path, rows)
        ok = "✓" if total_bytes == fc["byte_total"] else f"WARN got {total_bytes} expected {fc['byte_total']}"
        summary.append((key, n, total_bytes, fc["byte_total"], ok))
        print(f"  {key:<8} {n:>4} fields  {total_bytes:>4} bytes  expected {fc['byte_total']:>4}  {ok}")

    # State codes (only annual XLSX has a 'State code' sheet — calendar layouts
    # may or may not).
    state_csv = ROOT / "codemaps" / "state.csv"
    if not state_csv.exists():
        if write_state_codes(wb, state_csv):
            print(f"  state codes -> {state_csv.relative_to(ROOT)}")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "releases",
        nargs="*",
        choices=list(RELEASES),
        help="Release(s) to build; default: all known releases",
    )
    args = ap.parse_args()
    targets = args.releases or list(RELEASES)
    for r in targets:
        build_release(r)


if __name__ == "__main__":
    main()
