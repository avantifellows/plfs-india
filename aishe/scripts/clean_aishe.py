"""
Parse the AISHE Final Report workbooks in raw/ into the single denormalized
higher-ed fact (clean/higher_ed.parquet → BQ aishe_fact_higher_ed).

The fact unifies several published cuts into one grain. Every row is tagged
with `cut` (which published slice it came from) and `metric` (enrolment vs
graduates). Dimensions a given cut doesn't break out carry the sentinel "All".
ALWAYS filter on `cut`; the cuts overlap, so SUMMing across them double-counts.

  cut='state_level'      Table 33      graduates by state x level (2021-22)
  cut='programme_social' Table 34a     graduates by programme x social category (2021-22)
  cut='ug_discipline'    Tables 12 + 35  UG enrolment (T12) + graduates (T35) by
                                        discipline, 2019-20 -> 2021-22 (the trend)

metric='enrolment' exists only on the ug_discipline cut (from Table 12);
metric='graduates' exists on all three cuts.

Grain: (cut, aishe_year, metric, level, state, discipline, programme,
        social_category, gender) -> value

Usage:
  python3 scripts/clean_aishe.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
from sources import CLEAN, LATEST_YEAR, REPORTS, SENTINEL, TABLES

COLUMNS = ["cut", "aishe_year", "metric", "level", "state", "discipline",
           "programme", "social_category", "gender", "value"]

LEVELS = [
    "Ph.D.", "M.Phil.", "Post Graduate", "Under Graduate",
    "PG Diploma", "Diploma", "Certificate", "Integrated",
]
GENDERS = ["Male", "Female", "Total"]
SOCIAL_CATEGORIES = [
    "All Categories", "Scheduled Caste", "Scheduled Tribe",
    "Other Backward Classes", "Persons with Disability", "Muslim",
    "Other Minority Communities", "EWS",
]


def _wb(year: str):
    path = REPORTS[year]
    if not path.exists():
        raise SystemExit(
            f"missing raw workbook: {path}\n"
            f"Download the AISHE {year} Final Report and place it there (see README)."
        )
    return openpyxl.load_workbook(path, data_only=True)


def _sheet(wb, *names):
    want = {n.replace(" ", "").lower() for n in names}
    for s in wb.sheetnames:
        if s.replace(" ", "").lower() in want:
            return wb[s]
    raise SystemExit(f"no sheet matching {names} (have: {wb.sheetnames})")


def _row(cut, year, metric, level, state, discipline, programme,
         social_category, gender, value):
    return {
        "cut": cut, "aishe_year": year, "metric": metric, "level": level,
        "state": state, "discipline": discipline, "programme": programme,
        "social_category": social_category, "gender": gender,
        "value": int(value) if isinstance(value, (int, float)) else 0,
    }


# ─── Table 33: graduates by state × level (2021-22) ──────────────────────────
def state_level_rows(wb) -> list[dict]:
    ws = _sheet(wb, "33OutTurnState")
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        state = row[1]
        if state is None or not str(state).strip():
            continue
        state = str(state).strip()
        if state.lower() in {"all india", "india", "total"}:
            continue
        for li, level in enumerate(LEVELS):
            for gi, gender in enumerate(GENDERS):
                out.append(_row("state_level", LATEST_YEAR, "graduates", level,
                                state, SENTINEL, SENTINEL, SENTINEL, gender,
                                row[2 + li * 3 + gi]))
    return out


# ─── Table 34a: graduates by programme × social category (2021-22, all levels) ─
def programme_social_rows(wb) -> list[dict]:
    ws = _sheet(wb, "34a")
    out = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        prog = row[1]
        if prog is None or not str(prog).strip():
            continue
        prog = str(prog).strip()
        for ci, cat in enumerate(SOCIAL_CATEGORIES):
            for gi, gender in enumerate(GENDERS):
                idx = 2 + ci * 3 + gi
                val = row[idx] if idx < len(row) else None
                out.append(_row("programme_social", LATEST_YEAR, "graduates",
                                SENTINEL, SENTINEL, SENTINEL, prog, cat, gender, val))
    return out


# ─── Tables 12 / 35: UG by discipline (2019-20 → 2021-22) ────────────────────
def _discipline_series(ws, year, metric) -> list[dict]:
    """UG-by-discipline layout (shared by 12UGDisc enrolment and 35UGDisc
    graduates); shifts across years (S.No. column added in 2021-22)."""
    schema = None
    for ri in range(1, 6):
        cells = [str(c.value).strip() if c.value is not None else "" for c in ws[ri]]
        if cells and cells[0] == "Discipline":
            schema = "old"
            break
        if len(cells) >= 2 and cells[1] == "Discipline":
            schema = "new"
            break
    if schema is None:
        raise SystemExit(f"could not detect schema in sheet {ws.title!r}")
    if schema == "old":
        col_disc, col_subj, col_m, col_f, col_t = 0, 1, 2, 3, 4
    else:
        col_disc, col_subj, col_m, col_f, col_t = 1, 2, 3, 4, 5

    out = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or len(r) <= col_t:
            continue
        disc, subj, male, female, total = r[col_disc], r[col_subj], r[col_m], r[col_f], r[col_t]
        if disc is None or not str(disc).strip():
            continue
        disc_s = str(disc).strip()
        if disc_s.isdigit():
            continue
        is_total = (subj is None or str(subj).strip() == "") or disc_s.endswith("Total")
        if not is_total:
            continue
        clean = disc_s[:-len("Total")].strip() if disc_s.endswith("Total") else disc_s
        if not clean or not isinstance(total, (int, float)):
            continue
        for gender, val in (("Male", male), ("Female", female), ("Total", total)):
            out.append(_row("ug_discipline", year, metric, "Under Graduate",
                            SENTINEL, clean, SENTINEL, SENTINEL, gender, val))
    return out


def discipline_rows() -> list[dict]:
    """UG graduates (Table 35) + UG enrolment (Table 12), per year."""
    out = []
    for year in REPORTS:
        wb = _wb(year)
        out += _discipline_series(_sheet(wb, "35UGDisc"), year, "graduates")
        out += _discipline_series(_sheet(wb, "12UGDisc"), year, "enrolment")
    return out


def main() -> None:
    wb2122 = _wb(LATEST_YEAR)
    rows = state_level_rows(wb2122) + programme_social_rows(wb2122) + discipline_rows()
    df = pd.DataFrame(rows, columns=COLUMNS)
    df["value"] = df["value"].astype("Int64")

    CLEAN.mkdir(parents=True, exist_ok=True)
    out = TABLES[0].local_path
    df.to_parquet(out, index=False, engine="pyarrow")

    print(f"AISHE → {out.name}: {len(df):,} rows")
    for cut in ("state_level", "programme_social", "ug_discipline"):
        sub = df[df.cut == cut]
        by_metric = ", ".join(f"{m}={ (sub.metric == m).sum():,}"
                              for m in ("graduates", "enrolment") if (sub.metric == m).any())
        print(f"  cut={cut:<17} {len(sub):>6,} rows  ({by_metric})")

    # Validation: 2021-22 UG GRADUATES reconciles across the state and discipline cuts.
    g = df[(df.metric == "graduates") & (df.gender == "Total")]
    ug_state = g[(g.cut == "state_level") & (g.aishe_year == LATEST_YEAR)
                 & (g.level == "Under Graduate")].value.sum()
    ug_disc = g[(g.cut == "ug_discipline") & (g.aishe_year == LATEST_YEAR)].value.sum()
    ok = ug_state == ug_disc == 7754223
    print(f"  2021-22 UG graduates: state-cut={ug_state:,}  discipline-cut={ug_disc:,}  "
          f"{'OK' if ok else 'CHECK'}")
    print("✓ done.")


if __name__ == "__main__":
    main()
